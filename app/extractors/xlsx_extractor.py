import io
import openpyxl
from app.extractors.base import BaseExtractor, format_markdown_table
from app.models import ExtractionResult


class XLSXExtractor(BaseExtractor):
    async def extract(self, file_bytes: bytes, filename: str, **kwargs) -> ExtractionResult:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        content_parts = []
        total_cells = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                # Filter out completely empty trailing rows
                if any(c is not None and str(c).strip() != "" for c in row):
                    cleaned = [str(c).strip() if c is not None else "" for c in row]
                    rows_data.append(cleaned)
                    total_cells += len(cleaned)

            if rows_data:
                sheet_md = [f"## Sheet: {sheet_name}"]
                headers = rows_data[0]
                rows = rows_data[1:] if len(rows_data) > 1 else []
                sheet_md.append(format_markdown_table(headers, rows))
                content_parts.append("\n\n".join(sheet_md))

        final_content = "\n\n---\n\n".join(content_parts)
        word_count = len(final_content.split())

        return ExtractionResult(
            content=final_content,
            images=[],
            metadata={
                "filename": filename,
                "sheets": wb.sheetnames,
                "total_cells": total_cells,
                "engine": "local",
            },
            word_count=word_count,
            detected_type="xlsx",
        )
